import openpyxl
import os
import textwrap
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw 
import argparse


parser = argparse.ArgumentParser()
parser.add_argument('-a', '--all',  action="store_true", help='stores output to an /all directory for easy importing')
args = parser.parse_args()




#Perhaps an enum would have served me better here...
COLUMN_NAME = 1
COLUMN_COST = 2
COLUMN_HEALTH = 3
COLUMN_SPEED = 4
COLUMN_DAMAGE = 5
COLUMN_TEXT = 6
COLUMN_COLOR = 7
COLUMN_TYPE = 8
COLUMN_GROUP = 9

COST_LOCATION = (40, 30)
TITLE_LOCATION = (95, 32)
CENTER_BAR_LOCATION = (40, 240)
HEALTH_LOCATION = (40, 280) #Also used as text location for Spells.
SPEED_LOCATION = (40, 300) #This is shifted for towers. These names could have been better...
DAMAGE_LOCATION = (40, 320)
TEXT_LOCATION = (40, 340)

TEXT_RESOLUTION = 20 #Vertical distance between lines
SHEET_WIDTH = 26


TEXT_COLOR = (0,0,0) #black

wb = openpyxl.load_workbook('Dogs Of War.xlsx')
sheet = wb.get_sheet_by_name('cards')

titleFont = ImageFont.truetype("VerilySerifMono.otf", 16)
costFont = ImageFont.truetype("VerilySerifMono.otf", 32)
textFont = ImageFont.truetype("VerilySerifMono.otf", 16)
textFontLong = ImageFont.truetype("VerilySerifMono.otf", 12)


#Make all folder, if necessary
if args.all:
	if not os.path.exists('all'):
		os.makedirs('all')
		
		

for i in range(2,sheet.max_row + 1):
	folder = sheet.cell(row=i, column=COLUMN_COLOR).value
	subfolder = sheet.cell(row=i, column=COLUMN_TYPE).value
	
	if not os.path.exists(folder):
		os.makedirs(folder)
		
	if not os.path.exists(folder + '/' + subfolder):
		os.makedirs(folder + '/' + subfolder)
	
	#Get the right background
	img = Image.open(folder + '/' + folder + ".png")
	draw = ImageDraw.Draw(img)
	
	#Draw the title
	draw.text(TITLE_LOCATION, str(sheet.cell(row=i,column=COLUMN_NAME).value), TEXT_COLOR, font=titleFont)
	
	#Draw the center bar
	draw.text(CENTER_BAR_LOCATION, sheet.cell(row=i,column=COLUMN_TYPE).value + ', ' + sheet.cell(row=i,column=COLUMN_COLOR).value + ' ' + sheet.cell(row=i,column=COLUMN_GROUP).value, TEXT_COLOR, font=titleFont)

	#Draw cost
	draw.text(COST_LOCATION, str(sheet.cell(row=i,column=COLUMN_COST).value), TEXT_COLOR, font=costFont)
	
	#Draw type-specific things
	
	#Minion
	if sheet.cell(row=i,column=COLUMN_TYPE).value == "Minion":
		draw.text(HEALTH_LOCATION, "Health: " + str(sheet.cell(row=i,column=COLUMN_HEALTH).value), TEXT_COLOR, font=textFont)
		draw.text(SPEED_LOCATION, "Speed: " + str(sheet.cell(row=i,column=COLUMN_SPEED).value), TEXT_COLOR, font=textFont)
		draw.text(DAMAGE_LOCATION, "Damage: " + str(sheet.cell(row=i,column=COLUMN_DAMAGE).value), TEXT_COLOR, font=textFont)
		
		#Text wrap
		if sheet.cell(row=i,column=COLUMN_TEXT).value is not None:
			text = textwrap.wrap(str(sheet.cell(row=i,column=COLUMN_TEXT).value), width=SHEET_WIDTH)
			count = 0
			for t in text:
				draw.text((TEXT_LOCATION[0], TEXT_LOCATION[1] + count * TEXT_RESOLUTION), str(t), TEXT_COLOR, font=textFont)
				count = count + 1
		
	#Tower
	if sheet.cell(row=i,column=COLUMN_TYPE).value == "Tower":
		draw.text(HEALTH_LOCATION, "Attacks per Round: " + str(sheet.cell(row=i,column=COLUMN_SPEED).value), TEXT_COLOR, font=textFont)
		draw.text(SPEED_LOCATION, "Damage per Attack: " + str(sheet.cell(row=i,column=COLUMN_DAMAGE).value), TEXT_COLOR, font=textFont)
		
		#Text wrap
		if sheet.cell(row=i,column=COLUMN_TEXT).value is not None:
			text = textwrap.wrap(str(sheet.cell(row=i,column=COLUMN_TEXT).value), width=SHEET_WIDTH)
			myfont = textFont
			if len(text) > 4:#Max rows that fit on a sheet (after tower info)
				myfont = textFontLong
				text = textwrap.wrap(sheet.cell(row=i,column=COLUMN_TEXT).value, width=SHEET_WIDTH + 10)
			
			count = 0
			for t in text:
				draw.text((TEXT_LOCATION[0], TEXT_LOCATION[1] + count * TEXT_RESOLUTION), str(t), TEXT_COLOR, font=myfont)
				count = count + 1
			
	#Spell
	if sheet.cell(row=i,column=COLUMN_TYPE).value == "Spell":
		if sheet.cell(row=i,column=COLUMN_TEXT).value is not None:
			text = textwrap.wrap(sheet.cell(row=i,column=COLUMN_TEXT).value, width=SHEET_WIDTH)
			myfont = textFont
			if len(text) > 8:#Max rows that fit on a sheet
				myfont = textFontLong#Could make this a sanitize_text function
				text = textwrap.wrap(sheet.cell(row=i,column=COLUMN_TEXT).value, width=SHEET_WIDTH + 10)
			
			count = 0
			for t in text:
				draw.text((HEALTH_LOCATION[0], HEALTH_LOCATION[1] + count * TEXT_RESOLUTION), str(t), TEXT_COLOR, font=myfont)
				count = count + 1
				
			
	img.save(folder + '/' + subfolder + '/' + sheet.cell(row=i,column=COLUMN_NAME).value.replace(" ", "_") + '.png')
	
	#make a second copy for easy import
	if args.all:
		img.save('all' + '/' + sheet.cell(row=i,column=COLUMN_NAME).value.replace(" ", "_") + '.png')
		
	img.close()